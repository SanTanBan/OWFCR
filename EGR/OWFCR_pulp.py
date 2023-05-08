import math
from math import sqrt
from pathlib import Path
from time import time

import matplotlib.pylab as plt
import numpy as np
import openpyxl
import pandas as pd
from matplotlib.lines import Line2D
from pulp import (PULP_CBC_CMD, LpMinimize, LpProblem, LpStatus, LpVariable,
                  lpSum, value)
from scipy.spatial import Voronoi

handles, labels = plt.gca().get_legend_handles_labels()

colourSet = [
    'blue', 'orange', 'green', 'red', 'purple', 'brown', 'pink', 'grey',
    'olive', 'cyan'
]

Excel_Filename = "Dashboard.xlsx"

DashBoard = pd.read_excel(Excel_Filename, "DashBoard", index_col=0)

# Create a dictionary from the index and a specific column
Decisions = DashBoard['Responde'].to_dict()

# An additional technical requirement is that a given maximum number of cables, say C, can be connected to each substation
C = int(Decisions['C'])
TimeLimit = int(Decisions['TimeLimit'])
OWFCR_Type = int(Decisions['OWFCR_Type'])

if OWFCR_Type == 2:
    d_max = int(Decisions['d_max'])
    D = set(range(1, d_max + 1))
    pi_2 = int(Decisions['pi_2'])

    pi_OWFCR_BR_Cost = {}
    pi_OWFCR_BR_Cost[
        1] = 0  # No extra Cost as 1 incoming cable is already available for each Turbine
    pi_OWFCR_BR_Cost[2] = pi_2

elif OWFCR_Type == 4:
    Tao = "Export_Cable_Special"
    mu_1 = int(Decisions['mu_1'])
    mu_2 = int(Decisions['mu_2'])
    c_otm = int(Decisions['c_otm'])

No_Cross_Constraints = int(Decisions['No_Cross_Constraints'])
Common_No_Cross_Boundary_Radius = float(
    Decisions['Common_No_Cross_Boundary_Radius'])


def Euclidean_Distance_Matrix_Generator(locations_dict, Obstacles_Index_Set,
                                        Obstacles_Line_Start_Lat_Dict,
                                        Obstacles_Line_Start_Lon_Dict,
                                        Obstacles_Line_End_Lat_Dict,
                                        Obstacles_Line_End_Lon_Dict,
                                        output_filename):
    Distance_Dict = {}
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Euclidean Distances")

    # Write headers
    ws.cell(row=1, column=1, value="Location")
    for i, loc in enumerate(locations_dict):
        #print(i,loc)
        ws.cell(row=1, column=i + 2, value=loc)

    # Write distance matrix
    for i, loc1 in enumerate(locations_dict):
        ws.cell(row=i + 2, column=1, value=loc1)
        for j, loc2 in enumerate(locations_dict):
            if i == j:
                ws.cell(row=i + 2, column=j + 2, value=0)
                Distance_Dict[(loc1, loc2)] = 0
            elif j < i:
                ws.cell(row=i + 2,
                        column=j + 2,
                        value=ws.cell(row=j + 2, column=i + 2).value)
                Distance_Dict[(loc1, loc2)] = Distance_Dict[(loc2, loc1)]
            else:

                lat1, lon1, pow1 = locations_dict[loc1]
                lat2, lon2, pow2 = locations_dict[loc2]
                max_dist = sqrt((lat1 - lat2)**2 + (lon1 - lon2)**2)

                for individual_obstacle in Obstacles_Index_Set:
                    if do_lines_intersect(
                            locations_dict[loc1], locations_dict[loc2],
                        (Obstacles_Line_Start_Lat_Dict[individual_obstacle],
                         Obstacles_Line_Start_Lon_Dict[individual_obstacle],
                         0),
                        (Obstacles_Line_End_Lat_Dict[individual_obstacle],
                         Obstacles_Line_End_Lon_Dict[individual_obstacle], 0)):
                        distance = 9999999999999  # Very Large Value

                        if distance > max_dist:
                            max_dist = distance  # max_dist is the maximum distance between two points since the same two points would have the Euclidean distance without an Obstacle and will have the 999999999 distance with another obstacle

                ws.cell(row=i + 2, column=j + 2, value=max_dist)
                Distance_Dict[(loc1, loc2)] = max_dist
                if max_dist > 9999999999999:
                    for i in range(100):
                        print(
                            "Fatal Error!: Increase the Big M value by adding more 9's after 9999999999999 and also decrease all distances by a certain factor or power of 10"
                        )

    del wb["Sheet"]
    # Save workbook to file
    wb.save(output_filename)
    return Distance_Dict


def do_lines_intersect(line1_start, line1_end, line2_start, line2_end):
    """
    Check if two line segments intersect.

    Args:
        line1_start (tuple): The starting point of the first line segment, represented as a tuple of (x, y) coordinates.
        line1_end (tuple): The ending point of the first line segment, represented as a tuple of (x, y) coordinates.
        line2_start (tuple): The starting point of the second line segment, represented as a tuple of (x, y) coordinates.
        line2_end (tuple): The ending point of the second line segment, represented as a tuple of (x, y) coordinates.

    Returns:
        bool: True if the two line segments intersect, False otherwise.
    """
    x1, y1, p1 = line1_start
    x2, y2, p2 = line1_end
    x3, y3, p3 = line2_start
    x4, y4, p4 = line2_end

    # Calculate the slopes of the two lines
    slope1 = (y2 - y1) / (x2 - x1) if x2 - x1 != 0 else float('inf')
    slope2 = (y4 - y3) / (x4 - x3) if x4 - x3 != 0 else float('inf')

    # Check if the lines are parallel
    if slope1 == slope2:
        return False

    # Calculate the y-intercepts of the two lines
    intercept1 = y1 - slope1 * x1
    intercept2 = y3 - slope2 * x3

    # Calculate the x-coordinate of the point of intersection
    x_intersect = (intercept2 - intercept1) / (slope1 - slope2)

    # Check if the x-coordinate of the point of intersection is within the bounds of both line segments
    if x_intersect < min(x1, x2) or x_intersect > max(
            x1, x2) or x_intersect < min(x3, x4) or x_intersect > max(x3, x4):
        return False

    return True


def Intersecting_Arcs_Combo(locations_dict):
    Crossings_Set = set()
    for i, start_line1 in enumerate(locations_dict):
        for j, start_line2 in enumerate(locations_dict):
            #if start_line1!=start_line2 and i<j:
            if i < j:
                for k, end_line1 in enumerate(locations_dict):
                    #if end_line1!=start_line2 and start_line1!=end_line1 and i<k:
                    if end_line1 != start_line2 and i < k:
                        for l, end_line2 in enumerate(locations_dict):
                            #if end_line2!=end_line1 and end_line2!=start_line2 and end_line2!=start_line1 and j<l:
                            if end_line2 != end_line1 and j < l:
                                if do_lines_intersect(
                                        locations_dict[start_line1],
                                        locations_dict[end_line1],
                                        locations_dict[start_line2],
                                        locations_dict[end_line2]):
                                    #Crossings_Set.add(frozenset([frozenset([start_line1, end_line1]),frozenset([start_line2, end_line2])]))
                                    Crossings_Set.add(
                                        ((start_line1, end_line1),
                                         (start_line2, end_line2)))
                                else:
                                    continue
    return Crossings_Set


def Improved_Intersecting_Arcs_given_Line_Point(locations_dict,
                                                output_filename):
    Clique_Arc_Subset_Dict = {}
    for line_start in locations_dict:
        for line_end in locations_dict:
            #if start_line1!=start_line2 and i<j:
            if line_start != line_end:
                for clique_point in locations_dict:
                    if clique_point != line_start and clique_point != line_end:
                        value = [(line_start, line_end),
                                 (line_end, line_start)]
                        for query_point in locations_dict:
                            if query_point != line_start and query_point != line_end and query_point != clique_point:
                                if do_lines_intersect(
                                        locations_dict[line_start],
                                        locations_dict[line_end],
                                        locations_dict[clique_point],
                                        locations_dict[query_point]):
                                    value.append((clique_point, query_point))
                                    #value.append((query_point,clique_point))
                        Clique_Arc_Subset_Dict[line_start, line_end,
                                               clique_point] = value
    return Clique_Arc_Subset_Dict


def Intersecting_Arcs_within_Radius(locations_dict, Distance_Matrix,
                                    Radial_Crossings_Dict):
    Crossings_Set = set()
    for i, start_line1 in enumerate(locations_dict):
        for j, start_line2 in enumerate(locations_dict):
            #if start_line1!=start_line2 and i<j:
            if i < j and Distance_Matrix[
                (start_line1,
                 start_line2)] <= Radial_Crossings_Dict[start_line1]:
                for k, end_line1 in enumerate(locations_dict):
                    #if end_line1!=start_line2 and start_line1!=end_line1 and i<k:
                    if end_line1 != start_line2 and i < k and Distance_Matrix[
                        (start_line1,
                         end_line1)] <= Radial_Crossings_Dict[start_line1]:
                        for l, end_line2 in enumerate(locations_dict):
                            #if end_line2!=end_line1 and end_line2!=start_line2 and end_line2!=start_line1 and j<l:
                            if end_line2 != end_line1 and j < l and Distance_Matrix[
                                (start_line1, end_line2
                                 )] <= Radial_Crossings_Dict[start_line1]:
                                if do_lines_intersect(
                                        locations_dict[start_line1],
                                        locations_dict[end_line1],
                                        locations_dict[start_line2],
                                        locations_dict[end_line2]):
                                    #Crossings_Set.add(frozenset([frozenset([start_line1, end_line1]),frozenset([start_line2, end_line2])]))
                                    Crossings_Set.add(
                                        ((start_line1, end_line1),
                                         (start_line2, end_line2)))
                                else:
                                    continue
    return Crossings_Set


Turbines = pd.read_excel(Excel_Filename, "Turbine Locations", index_col=0)

Substations = pd.read_excel(Excel_Filename,
                            "Substation Locations",
                            index_col=0)
Steiners = pd.read_excel(Excel_Filename, "Steiner Locations", index_col=0)
Obstacles = pd.read_excel(Excel_Filename, "Obstacles", index_col=0)
Cables = pd.read_excel(Excel_Filename, "Cable Specifications", index_col=0)

Cables_Index_Set = set(Cables.index)
Cable_Cost_Dict = Cables["Cost"].to_dict()
Cable_Capacity_Dict = Cables["Capacity"].to_dict()

min_cable_cost = 9999999999999
for cable_type, cable_cost in Cable_Cost_Dict.items():
    if min_cable_cost > cable_cost:
        min_cable_cost = cable_cost
print("The Minimum Cable Cost is ", min_cable_cost)

Turbines_Index_Set = set(Turbines.index)
Turbines_Latitude_Dict = Turbines["Latitude"].to_dict()
Turbines_Longitude_Dict = Turbines["Longitude"].to_dict()
Turbines_Power_Dict = Turbines["Power"].to_dict()
Turbines_Crossing_Radius_Dict = Turbines[
    "Individual No-Cross Boundary Radius"].to_dict()

Substations_Index_Set = set(Substations.index)
Substations_Latitude_Dict = Substations["Latitude"].to_dict()
Substations_Longitude_Dict = Substations["Longitude"].to_dict()
Substations_Power_Dict = Substations["Power"].to_dict()
Substations_Crossing_Radius_Dict = Substations[
    "Individual No-Cross Boundary Radius"].to_dict()

Steiners_Index_Set = set(Steiners.index)
Steiners_Latitude_Dict = Steiners["Latitude"].to_dict()
Steiners_Longitude_Dict = Steiners["Longitude"].to_dict()
Steiners_Power_Dict = Steiners["Power"].to_dict()
Steiners_Crossing_Radius_Dict = Steiners[
    "Individual No-Cross Boundary Radius"].to_dict()

Obstacles_Index_Set = set(Obstacles.index)
Obstacles_Line_Start_Lat_Dict = Obstacles["Line Start Latitude"].to_dict()
Obstacles_Line_Start_Lon_Dict = Obstacles["Line Start Longitude"].to_dict()
Obstacles_Line_End_Lat_Dict = Obstacles["Line End Latitude"].to_dict()
Obstacles_Line_End_Lon_Dict = Obstacles["Line End Longitude"].to_dict()

All_Nodes_Index_Set = Turbines_Index_Set | Substations_Index_Set | Steiners_Index_Set
All_Nodes_DictKey_Lat_Lon_Power_ValueTuple = {}
for Single_Node in All_Nodes_Index_Set:
    if Single_Node in Turbines_Index_Set:
        ValueTuple = (Turbines_Latitude_Dict[Single_Node],
                      Turbines_Longitude_Dict[Single_Node],
                      Turbines_Power_Dict[Single_Node])
    elif Single_Node in Substations_Index_Set:
        ValueTuple = (Substations_Latitude_Dict[Single_Node],
                      Substations_Longitude_Dict[Single_Node],
                      Substations_Power_Dict[Single_Node])
    elif Single_Node in Steiners_Index_Set:
        ValueTuple = (Steiners_Latitude_Dict[Single_Node],
                      Steiners_Longitude_Dict[Single_Node],
                      Steiners_Power_Dict[Single_Node])
    else:
        print("Fatal Error! ", Single_Node, " not found anywhere")
    All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[Single_Node] = ValueTuple

Distance_Dict = Euclidean_Distance_Matrix_Generator(
    All_Nodes_DictKey_Lat_Lon_Power_ValueTuple, Obstacles_Index_Set,
    Obstacles_Line_Start_Lat_Dict, Obstacles_Line_Start_Lon_Dict,
    Obstacles_Line_End_Lat_Dict, Obstacles_Line_End_Lon_Dict,
    "Distance Matrix.xlsx")
Crossing_Radius_Dict = {
    **Steiners_Crossing_Radius_Dict,
    **Substations_Crossing_Radius_Dict,
    **Turbines_Crossing_Radius_Dict
}
if No_Cross_Constraints == 0:

    if Common_No_Cross_Boundary_Radius <= 0:
        Crossings_Set = Intersecting_Arcs_Combo(
            All_Nodes_DictKey_Lat_Lon_Power_ValueTuple)
    else:
        Crossings_Set = Intersecting_Arcs_within_Radius(
            All_Nodes_DictKey_Lat_Lon_Power_ValueTuple, Distance_Dict,
            Crossing_Radius_Dict)

    print("Number of Crossings found: ", len(Crossings_Set))

elif No_Cross_Constraints == 1:
    Clique_Arc_Subset_Dict = Improved_Intersecting_Arcs_given_Line_Point(
        All_Nodes_DictKey_Lat_Lon_Power_ValueTuple,
        "Crossing Segments New Eq 14.xlsx")
    print("Length of the Dictionary: ", len(Clique_Arc_Subset_Dict))

# Formulating the Problem
prob = LpProblem("OWFCR_Tree_Type_" + str(OWFCR_Type), LpMinimize)

all_connections = [(i, j) for i in All_Nodes_Index_Set
                   for j in All_Nodes_Index_Set if i != j]


def zip_coords(df):
    return list(zip(df["Latitude"], df["Longitude"]))


v = Voronoi(np.array(zip_coords(Turbines)))

indices = list(Turbines.index) + list(Substations.index) + list(Steiners.index)
substations_steiners = list(Substations.index) + list(Steiners.index)


def two_way_conn(conns):
    return [*((i, j) for (i, j) in conns), *((j, i) for (i, j) in conns)]


connections = list(
    set(
        two_way_conn([(indices[a[0]], indices[a[1]])
                      for a in v.ridge_points.tolist()] +
                     [(i, j) for i in indices
                      for j in substations_steiners if i != j])))
cables_per_connections = [(i, j, t) for (i, j) in connections
                          for t in Cables_Index_Set]

# Decision Variables

f = LpVariable.dicts('f', connections, lowBound=0)

x = LpVariable.dicts('x', cables_per_connections, cat='Binary')

y = LpVariable.dicts('y', connections, cat='Binary')

if OWFCR_Type == 2:
    z = LpVariable.dicts('z', ((j, d) for j in Turbines_Index_Set for d in D),
                         cat='Binary')

if OWFCR_Type == 3:
    q = LpVariable.dicts('q', connections, cat='Binary')

# Minimization Objective Function
if OWFCR_Type == 0 or OWFCR_Type == 1:  # Equation 1 for OWFCR or SS
    prob += lpSum(Cable_Cost_Dict[t] * Distance_Dict[(i, j)] * x[i, j, t]
                  for (i, j, t) in cables_per_connections)
    #
elif OWFCR_Type == 2:  # Equation 16 for BP
    prob += lpSum(Cable_Cost_Dict[t] * Distance_Dict[(i, j)] * x[i, j, t]
                  for (i, j, t) in cables_per_connections) + lpSum(
                      pi_OWFCR_BR_Cost[d] * z[j, d] for d in D
                      for j in Turbines_Index_Set)

elif OWFCR_Type == 3:  # Equation 24 for CL
    prob += lpSum(Cable_Cost_Dict[t] * Distance_Dict[(i, j)] * x[i, j, t]
                  for (i, j, t) in cables_per_connections) + lpSum(
                      min_cable_cost * Distance_Dict[(i, j)] * q[i, j]
                      for (i, j) in connections)

elif OWFCR_Type == 4:  # Equation 27 for OTM
    prob += lpSum(Cable_Cost_Dict[t] * Distance_Dict[(i, j)] * x[i, j, t]
                  for (i, j, t) in cables_per_connections) + lpSum(
                      c_otm * x[i, j, Tao] for (i, j) in connections)

# Equation 2
for (i, j) in connections:
    prob += lpSum(x[i, j, t] for t in Cables_Index_Set) == y[i, j]


def connected(h):
    return [i for i in All_Nodes_Index_Set if (i, h) in connections]


# Equation 3
for h in (Turbines_Index_Set | Steiners_Index_Set):
    prob += lpSum(f[h, i] - f[i, h] for i in connected(
        h)) == All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[h][2]

# Equation 4
for (i, j) in connections:
    prob += lpSum(Cable_Capacity_Dict[t] * x[i, j, t]
                  for t in Cables_Index_Set) >= f[i, j]

# Equation 5
for h in Turbines_Index_Set:
    prob += lpSum(y[h, j] for j in connected(h)) == 1

# Equation 6
for h in Substations_Index_Set:
    prob += lpSum(y[h, j] for j in connected(h)) == 0

# Equation 7
for h in Steiners_Index_Set:
    prob += lpSum(y[h, j] for j in connected(h)) <= 1

# Equation 8
for h in Steiners_Index_Set:
    prob += lpSum(y[i, h] for i in connected(h)) <= 1

# Equation 9
for h in Substations_Index_Set:
    prob += lpSum(y[i, h] for i in connected(h)) <= C

if No_Cross_Constraints == 0:
    # Equation 10
    for Each_Crossing in Crossings_Set:
        Line1, Line2 = Each_Crossing
        h, k = Line1
        i, j = Line2
        if (i, j) in connections and (h, k) in connections:
            prob += y[i, j] + y[j, i] + y[h, k] + y[k, h] <= 1
elif No_Cross_Constraints == 1:
    # Equation 14
    for (a, b) in connections:
        for f in All_Nodes_Index_Set:
            if f != a and f != b:
                prob += lpSum(
                    y[i, j] for i, j in Clique_Arc_Subset_Dict[(a, b, f)]) <= 1

# Equation 15
if OWFCR_Type == 1:
    for h in Turbines_Index_Set:
        prob += lpSum(y[i, h] for i in connected(h)) <= 1

# Equation 17 and 18
if OWFCR_Type == 2:
    for j in Turbines_Index_Set:
        prob += lpSum(y[i, j] for i in connected(j)) == lpSum(d * z[j, d]
                                                              for d in D)

        prob += lpSum(z[j, d] for d in D) <= 1

#Equations 20, 21 and 23
if OWFCR_Type == 3:
    for h in (Turbines_Index_Set | Steiners_Index_Set):
        prob += lpSum(
            y[i, h] + y[h, i] + q[i, h] + q[h, i]
            for i in connected(h)) == 2 * lpSum(y[h, j] for j in connected(h))

    for (i, j) in connections:
        if i > j:
            prob += q[i, j] == 0

    for Each_Crossing in Crossings_Set:
        Line1, Line2 = Each_Crossing
        h, k = Line1
        i, j = Line2
        prob += y[i, j] + y[j, i] + y[h, k] + y[k, h] <= 1
        prob += y[i, j] + y[j, i] + y[h, k] + y[k, h] + q[i, j] + q[j, i] + q[
            h, k] + q[k, h] <= 1

# Equation 25 and 26
if OWFCR_Type == 4:
    for h in Turbines_Index_Set:
        prob += lpSum(x[i, h, t] for i in connected(h)
                      for t in Cables_Index_Set if t != Tao) <= mu_1

    for h in Turbines_Index_Set:
        prob += lpSum(x[i, h, Tao] for i in connected(h)) <= mu_2

# Solve the Problem using default PuLP_CBC settings
start_time = time()

from multiprocessing import cpu_count

seed = 1234567
status = prob.solve(
    PULP_CBC_CMD(timeLimit=TimeLimit,
                 threads=cpu_count(),
                 options=[f"RandomS {seed}"]))

end_time = time()

run_time = end_time - start_time
print("This is the status:- ", LpStatus[prob.status])
objec_val = value(prob.objective)
print("Objective : ", objec_val, " found in ", run_time, " seconds")

if OWFCR_Type == 3:
    rings_or_closed_loops = []
    for t in Cables_Index_Set:
        if Cable_Cost_Dict[t] == min_cable_cost:
            for i in All_Nodes_Index_Set:
                for j in All_Nodes_Index_Set:
                    if i != j:
                        if value(q[i, j]) == 1:
                            rings_or_closed_loops.append((i, j))
            break

All_Used_Steiner_Nodes = set()
Cable_Routes_Dict = {}

# Plotting the Turbines, Substations, Steiner Nodes and Cable Routes in Separate Cable-Wise Map
for t in Cables_Index_Set:
    routes = []
    Used_Steiner_NodeSet = set()
    for (i, j) in connections:
        if value(x[i, j, t]) == 1:
            routes.append((i, j))
            if i in Steiners_Index_Set:
                Used_Steiner_NodeSet.add(i)
            elif j in Steiners_Index_Set:
                Used_Steiner_NodeSet.add(j)

    All_Used_Steiner_Nodes.union(Used_Steiner_NodeSet)
    Cable_Routes_Dict[t] = routes

# Plotting the Turbines, Substations, Steiner Nodes and Cable Routes in Single Diagram
plt.figure(figsize=(13, 13))

All_Used_Steiner_Nodes = Steiners_Index_Set
Steiner_Handle = plt.scatter([
    All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1]
    for i in All_Used_Steiner_Nodes
], [
    All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]
    for i in All_Used_Steiner_Nodes
],
                             c='blue',
                             marker="*",
                             label="Steiner Nodes")
handles.extend([Steiner_Handle])
for i in All_Used_Steiner_Nodes:
    plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33,
             All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)

# plot all turbines with the same marker and color, and assign the label outside the loop
Turbine_Handle = plt.scatter([
    All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1]
    for i in Turbines_Index_Set
], [
    All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]
    for i in Turbines_Index_Set
],
                             c='r',
                             marker='s',
                             label="Turbines")
handles.extend([Turbine_Handle])
for i in Turbines_Index_Set:
    plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33,
             All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)

Substation_Handle = plt.scatter([
    All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1]
    for i in Substations_Index_Set
], [
    All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]
    for i in Substations_Index_Set
],
                                c='black',
                                label="Substations")
handles.extend([Substation_Handle])
for i in Substations_Index_Set:
    plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33,
             All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)

min_line_capacity = min(*(c for c in Cable_Capacity_Dict.values()))

for t in Cables.index:
    if t not in Cable_Routes_Dict:
        continue

    routes = Cable_Routes_Dict[t]
    Colour_Me = colourSet[list(Cables.index).index(t)]
    dynamic_line_width = 2 * math.log(Cable_Capacity_Dict[t] /
                                      min_line_capacity + 1)

    if OWFCR_Type == 3:
        if Cable_Cost_Dict[t] == min_cable_cost:
            for i, j in rings_or_closed_loops:
                plt.annotate(
                    '',
                    xy=[
                        All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1],
                        All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]
                    ],
                    xytext=[
                        All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1],
                        All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]
                    ],
                    arrowprops=dict(arrowstyle="<|-|>",
                                    connectionstyle='arc3',
                                    edgecolor=Colour_Me,
                                    linewidth=dynamic_line_width))
                # https://matplotlib.org/stable/gallery/text_labels_and_annotations/fancyarrow_demo.html

    for i, j in routes:
        plt.annotate('',
                     xy=[
                         All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1],
                         All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]
                     ],
                     xytext=[
                         All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1],
                         All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]
                     ],
                     arrowprops=dict(arrowstyle="simple",
                                     connectionstyle='arc3',
                                     edgecolor=Colour_Me,
                                     linewidth=dynamic_line_width))

    line_handle = Line2D([0], [0], label=t, color=Colour_Me)
    handles.extend([line_handle])

for individual_obstacle in Obstacles_Index_Set:
    obstacle_Lon, obstacle_Lat = [
        Obstacles_Line_Start_Lon_Dict[individual_obstacle],
        Obstacles_Line_End_Lon_Dict[individual_obstacle]
    ], [
        Obstacles_Line_Start_Lat_Dict[individual_obstacle],
        Obstacles_Line_End_Lat_Dict[individual_obstacle]
    ]
    plt.plot(obstacle_Lon, obstacle_Lat, linestyle="dashdot", color="black")
    mid_Lat = (Obstacles_Line_Start_Lat_Dict[individual_obstacle] +
               Obstacles_Line_End_Lat_Dict[individual_obstacle]) / 2
    mid_Lon = (Obstacles_Line_Start_Lon_Dict[individual_obstacle] +
               Obstacles_Line_End_Lon_Dict[individual_obstacle]) / 2
    plt.text(mid_Lon, mid_Lat, individual_obstacle)

line_handle = Line2D([0], [0],
                     linestyle="dashdot",
                     label="Obstacles",
                     color="black")
handles.extend([line_handle])

if No_Cross_Constraints == 0:
    no_cross = " & Eq 10"
elif No_Cross_Constraints == 1:
    no_cross = " & Eq 14"
else:
    no_cross = " without no-cross-constraints"
solver_name = "PuLP_CBC" + no_cross

model_name = {
    0: "OWFCR",
    1: "OWFCR_SS",
    2: "OWFCR_BP",
    3: "OWFCR_CL",
    4: "OWFCR_OTM"
}[OWFCR_Type]

name = (f"{model_name} Objective Value {objec_val:.0f} kEuro found in "
        f"{run_time:.0f} seconds for {len(Turbines_Index_Set)} Turbines "
        f"using {solver_name}")

plt.title(name)
plt.ylabel("Latitude")
plt.xlabel("Longitude")
plt.legend(handles=handles)
#plt.legend()
Path("images").mkdir(exist_ok=True)
name = "images/" + name + ".png"
plt.savefig(name, format='png', bbox_inches="tight")
