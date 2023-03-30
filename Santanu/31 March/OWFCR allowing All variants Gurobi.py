import random
import winsound
import openpyxl
from math import sqrt
import matplotlib.pylab as plt
from matplotlib.lines import Line2D
import pandas as pd
from gurobipy import *
from time import time

handles, labels = plt.gca().get_legend_handles_labels()
tolerance=1e-5 # This is the default value for Gurobi
#colourSet=['b','g','r','c','m','y','blue','orange','green','red','purple','brown','pink','grey','olive','cyan']
#colourSet=['b','g','r']
colourSet=['blue','orange','green','red','purple','brown','pink','grey','olive','cyan']

Excel_Filename="Dashboard.xlsx"

DashBoard=pd.read_excel(Excel_Filename,"DashBoard",index_col=0)

# Create a dictionary from the index and a specific column
Decisions = DashBoard['Responde'].to_dict()

# An additional technical requirement is that a given maximum number of cables, say C, can be connected to each substation
C = int(Decisions['C'])

TimeLimit=int(Decisions['TimeLimit'])
No_Cross_Constraints=int(Decisions['No_Cross_Constraints'])
OWFCR_Type=int(Decisions['OWFCR_Type'])

if OWFCR_Type==2:
    d_max=int(Decisions['d_max'])
    D = set(range(1, d_max+1))
    pi_2=int(Decisions['pi_2'])

    pi_OWFCR_BR_Cost={}
    pi_OWFCR_BR_Cost[1]=0 # No extra Cost as 1 incoming cable is already available for each Turbine
    pi_OWFCR_BR_Cost[2]=pi_2

elif OWFCR_Type==4:
    Tao="Export_Cable_Special"
    mu_1=int(Decisions['mu_1'])
    mu_2=int(Decisions['mu_2'])
    c_otm=int(Decisions['c_otm'])



def do_lines_intersect(line1_start, line1_end, line2_start, line2_end):
    """
    Check if two line segments intersect.

    Args:
        line1_start (tuple): The starting point of the first line segment, represented as a tuple of (x, y) coordinates.
        line1_end (tuple): The ending point of the first line segment, represented as a tuple of (x, y) coordinates.
        line2_start (tuple): The starting point of the second line segment, represented as a tuple of (x, y) coordinates.
        line2_end (tuple): The ending point of the second line segment, represented as a tuple of (x, y) coordinates.

    Returns:
        bool: True if the two line segments intersect, False otherwise.
    """
    x1, y1 , p1 = line1_start
    x2, y2 , p2 = line1_end
    x3, y3 , p3 = line2_start
    x4, y4 , p4 = line2_end

    # Calculate the slopes of the two lines
    slope1 = (y2 - y1) / (x2 - x1) if x2 - x1 != 0 else float('inf')
    slope2 = (y4 - y3) / (x4 - x3) if x4 - x3 != 0 else float('inf')

    # Check if the lines are parallel
    if slope1 == slope2:
        return False
    elif x1-x2==0 and x3-x4==0:
        return False

    # Calculate the y-intercepts of the two lines
    intercept1 = y1 - slope1 * x1
    intercept2 = y3 - slope2 * x3

    # Calculate the x-coordinate of the point of intersection
    x_intersect = (intercept2 - intercept1) / (slope1 - slope2)

    # Check if the x-coordinate of the point of intersection is within the bounds of both line segments
    if x_intersect < min(x1, x2) or x_intersect > max(x1, x2) or x_intersect < min(x3, x4) or x_intersect > max(x3, x4):
        return False

    return True




# def Euclidean_Distance_Matrix_Generator(locations_dict, Obstacles_Index_Set, Obstacles_Line_Start_Lat_Dict, Obstacles_Line_Start_Lon_Dict, Obstacles_Line_End_Lat_Dict, Obstacles_Line_End_Lon_Dict, output_filename):
#     Distance_Dict = {}
#     # Create a new workbook
#     wb = openpyxl.Workbook()
#     ws = wb.create_sheet("Euclidean Distances")

#     # Write headers
#     ws.cell(row=1, column=1, value="Location")
#     for i, loc in enumerate(locations_dict):
#         #print(i,loc)
#         ws.cell(row=1, column=i+2, value=loc)
    
#     # Write distance matrix
#     for i, loc1 in enumerate(locations_dict):
#         ws.cell(row=i+2, column=1, value=loc1)
#         for j, loc2 in enumerate(locations_dict):
#             if i == j:
#                 ws.cell(row=i+2, column=j+2, value=0)
#                 Distance_Dict[(loc1,loc2)] = 0
#             elif j < i:
#                 ws.cell(row=i+2, column=j+2, value=ws.cell(row=j+2, column=i+2).value)
#                 Distance_Dict[(loc1,loc2)] = Distance_Dict[(loc2,loc1)]
#             else:
                
#                 for individual_obstacle in Obstacles_Index_Set:
#                     if do_lines_intersect(locations_dict[loc1],locations_dict[loc2],(Obstacles_Line_Start_Lat_Dict[individual_obstacle],Obstacles_Line_Start_Lon_Dict[individual_obstacle],0),(Obstacles_Line_End_Lat_Dict[individual_obstacle], Obstacles_Line_End_Lon_Dict[individual_obstacle],0)):
#                         distance=9999999999999 # Very Large Value
#                         ws.cell(row=i+2, column=j+2, value=distance)
#                         Distance_Dict[(loc1,loc2)] = distance
#                     else:                        
#                         lat1, lon1, pow1 = locations_dict[loc1]
#                         lat2, lon2, pow2 = locations_dict[loc2]

#                         #distance = round(sqrt((lat1-lat2)**2 + (lon1-lon2)**2),2)
#                         distance = sqrt((lat1-lat2)**2 + (lon1-lon2)**2)
#                         ws.cell(row=i+2, column=j+2, value=distance)
#                         Distance_Dict[(loc1,loc2)] = distance
    
#     del wb["Sheet"]
#     # Save workbook to file
#     wb.save(output_filename)
#     return Distance_Dict



def Euclidean_Distance_Matrix_Generator(locations_dict, Obstacles_Index_Set, Obstacles_Line_Start_Lat_Dict, Obstacles_Line_Start_Lon_Dict, Obstacles_Line_End_Lat_Dict, Obstacles_Line_End_Lon_Dict, output_filename):
    Distance_Dict = {}
    # Write distance matrix
    for loc1 in locations_dict:
        for loc2 in locations_dict:
            if loc1 == loc2:
                Distance_Dict[(loc1,loc2)] = 0
            else:
                for individual_obstacle in Obstacles_Index_Set:
                    if do_lines_intersect(locations_dict[loc1],locations_dict[loc2],(Obstacles_Line_Start_Lat_Dict[individual_obstacle],Obstacles_Line_Start_Lon_Dict[individual_obstacle],0),(Obstacles_Line_End_Lat_Dict[individual_obstacle], Obstacles_Line_End_Lon_Dict[individual_obstacle],0)):
                        distance=9999999999 # Very Large Value
                        Distance_Dict[(loc1,loc2)] = distance
                    else:                        
                        lat1, lon1, pow1 = locations_dict[loc1]
                        lat2, lon2, pow2 = locations_dict[loc2]
                        #distance = round(sqrt((lat1-lat2)**2 + (lon1-lon2)**2),2)
                        distance = sqrt((lat1-lat2)**2 + (lon1-lon2)**2)
                        Distance_Dict[(loc1,loc2)] = distance
    return Distance_Dict



def Intersecting_Arcs_Combo(locations_dict, output_filename):

    """wb = openpyxl.Workbook()
    ws = wb.create_sheet("Crossing Segments")
    row_count=1
    ws.cell(row=row_count, column=1, value="Line 1 Start")
    ws.cell(row=row_count, column=2, value="Line 1 End")
    ws.cell(row=row_count, column=3, value="Line 2 Start")
    ws.cell(row=row_count, column=4, value="Line 2 End")"""

    Crossings_Set=set()

    for i, start_line1 in enumerate(locations_dict):
        for j, start_line2 in enumerate(locations_dict):
            #if start_line1!=start_line2 and i<j:
            if i<j:
                for k, end_line1 in enumerate(locations_dict):
                    #if end_line1!=start_line2 and start_line1!=end_line1 and i<k:
                    if end_line1!=start_line2 and i<k:
                        for l, end_line2 in enumerate(locations_dict):
                            #if end_line2!=end_line1 and end_line2!=start_line2 and end_line2!=start_line1 and j<l:
                            if end_line2!=end_line1 and j<l:
                                if do_lines_intersect(locations_dict[start_line1], locations_dict[end_line1], locations_dict[start_line2], locations_dict[end_line2]):
                                    #Crossings_Set.add(frozenset([frozenset([start_line1, end_line1]),frozenset([start_line2, end_line2])]))
                                    Crossings_Set.add(((start_line1, end_line1),(start_line2, end_line2)))
                                    """row_count+=1
                                    ws.cell(row=row_count, column=1, value=start_line1)
                                    ws.cell(row=row_count, column=2, value=end_line1)
                                    ws.cell(row=row_count, column=3, value=start_line2)
                                    ws.cell(row=row_count, column=4, value=end_line2)
                                    # Commenting out the Crossing Creation; we need to actually create new sheets if row number exceeds 1048576"""
                                else:
                                    continue

    """del wb["Sheet"]
    # Save workbook to file
    wb.save(output_filename)"""
    return Crossings_Set



def Improved_Intersecting_Arcs_given_Line_Point(locations_dict, output_filename):
    Clique_Arc_Subset_Dict={}
    for line_start in locations_dict:
        for line_end in locations_dict:
            #if start_line1!=start_line2 and i<j:
            if line_start!=line_end:
                for clique_point in locations_dict:
                    if clique_point!=line_start and clique_point!=line_end:
                        value=[(line_start,line_end),(line_end,line_start)]
                        for query_point in locations_dict:
                            if query_point!=line_start and query_point!=line_end and query_point!=clique_point:
                                if do_lines_intersect(locations_dict[line_start], locations_dict[line_end], locations_dict[clique_point], locations_dict[query_point]):
                                    value.append((clique_point,query_point))
                                    #value.append((query_point,clique_point))
                        Clique_Arc_Subset_Dict[line_start,line_end,clique_point] = value
    return Clique_Arc_Subset_Dict



Turbines=pd.read_excel(Excel_Filename,"Turbine Locations",index_col=0)
Substations=pd.read_excel(Excel_Filename,"Substation Locations",index_col=0)
Steiners=pd.read_excel(Excel_Filename,"Steiner Locations",index_col=0)
Obstacles=pd.read_excel(Excel_Filename,"Obstacles",index_col=0)
Cables=pd.read_excel(Excel_Filename,"Cable Specifications",index_col=0)

# print(Turbines)
# print(Substations)
# print(Steiners)
# print(Cables)

Cables_Index_Set = set(Cables.index)
Cable_Cost_Dict = Cables["Cost"].to_dict()
Cable_Capacity_Dict = Cables["Capacity"].to_dict()

min_cable_cost=9999999999999
for cable_type,cable_cost in Cable_Cost_Dict.items():
    if min_cable_cost>cable_cost:
        min_cable_cost=cable_cost
print("The Minimum Cable Cost is " , min_cable_cost)

Turbines_Index_Set = set(Turbines.index)
Turbines_Latitude_Dict = Turbines["Latitude"].to_dict()
Turbines_Longitude_Dict = Turbines["Longitude"].to_dict()
Turbines_Power_Dict = Turbines["Power"].to_dict()

Substations_Index_Set = set(Substations.index)
Substations_Latitude_Dict = Substations["Latitude"].to_dict()
Substations_Longitude_Dict = Substations["Longitude"].to_dict()
Substations_Power_Dict = Substations["Power"].to_dict()

Steiners_Index_Set = set(Steiners.index)
Steiners_Latitude_Dict = Steiners["Latitude"].to_dict()
Steiners_Longitude_Dict = Steiners["Longitude"].to_dict()
Steiners_Power_Dict = Steiners["Power"].to_dict()

Obstacles_Index_Set = set(Obstacles.index)
Obstacles_Line_Start_Lat_Dict = Obstacles["Line Start Latitude"].to_dict()
Obstacles_Line_Start_Lon_Dict = Obstacles["Line Start Longitude"].to_dict()
Obstacles_Line_End_Lat_Dict = Obstacles["Line End Latitude"].to_dict()
Obstacles_Line_End_Lon_Dict = Obstacles["Line End Longitude"].to_dict()


All_Nodes_Index_Set = Turbines_Index_Set | Substations_Index_Set | Steiners_Index_Set
All_Nodes_DictKey_Lat_Lon_Power_ValueTuple={}
for Single_Node in All_Nodes_Index_Set:
    if Single_Node in Turbines_Index_Set:
        ValueTuple=(Turbines_Latitude_Dict[Single_Node],Turbines_Longitude_Dict[Single_Node],Turbines_Power_Dict[Single_Node])
    elif Single_Node in Substations_Index_Set:
        ValueTuple=(Substations_Latitude_Dict[Single_Node],Substations_Longitude_Dict[Single_Node],Substations_Power_Dict[Single_Node])
    elif Single_Node in Steiners_Index_Set:
        ValueTuple=(Steiners_Latitude_Dict[Single_Node],Steiners_Longitude_Dict[Single_Node],Steiners_Power_Dict[Single_Node])
    else:
        print("Fatal Error! ",Single_Node," not found anywhere")
    All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[Single_Node]=ValueTuple

# print(All_Nodes_Index_Set)
# print(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple)

Distance_Dict = Euclidean_Distance_Matrix_Generator(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple, Obstacles_Index_Set, Obstacles_Line_Start_Lat_Dict, Obstacles_Line_Start_Lon_Dict, Obstacles_Line_End_Lat_Dict, Obstacles_Line_End_Lon_Dict, "Distance Matrix.xlsx")
# print(Distance_Dict)

if No_Cross_Constraints==0:
    Crossings_Set=Intersecting_Arcs_Combo(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple, "Crossing Segments.xlsx")
    print("Number of Crossings found: ",len(Crossings_Set))

    # for Each_Crossing in Crossings_Set:
    #     Line1 , Line2 = Each_Crossing
    #     h , k = Line1
    #     i , j = Line2
    #     print(h,k,i,j)

elif No_Cross_Constraints == 1:
    Clique_Arc_Subset_Dict = Improved_Intersecting_Arcs_given_Line_Point(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple, "Crossing Segments New Eq 14.xlsx")
    print("Length of the Dictionary: ",len(Clique_Arc_Subset_Dict))

    # for a in All_Nodes_Index_Set:
    #     for b in All_Nodes_Index_Set:
    #         if a!=b:
    #             for f in All_Nodes_Index_Set:
    #                 if f!=a and f!=b:
    #                     print(Clique_Arc_Subset_Dict[(a,b,f)])

# Set the problem
mdl=Model("OWFCR_Tree_Type_"+str(OWFCR_Type))
mdl.setParam('TimeLimit', TimeLimit) # 3600 seconds is 1 hour

f, x, y = {}, {}, {}
if OWFCR_Type==2:
    z={}
if OWFCR_Type==3:
    q={}

# Decision Variables

for i in All_Nodes_Index_Set:
    for j in All_Nodes_Index_Set:
        if i!=j:
            f[i, j] = mdl.addVar(vtype=GRB.CONTINUOUS, name="x%s,%s" % (i, j))
            mdl.update()

for i in All_Nodes_Index_Set:
    for j in All_Nodes_Index_Set:
        if i!=j:
            for t in Cables_Index_Set:
                x[i, j, t] = mdl.addVar(vtype=GRB.BINARY, name="x%s,%s,%s" % (i, j, t))
                mdl.update()

for i in All_Nodes_Index_Set:
    for j in All_Nodes_Index_Set:
        if i!=j:
            y[i, j] = mdl.addVar(vtype=GRB.BINARY, name="x%s,%s" % (i, j))
            mdl.update()

if OWFCR_Type==2:
    for j in Turbines_Index_Set:
        for d in D:
            z[j, d] = mdl.addVar(vtype=GRB.BINARY, name="x%s,%s" % (j, d))
            mdl.update()

if OWFCR_Type==3:
    for i in All_Nodes_Index_Set:
        for j in All_Nodes_Index_Set:
            if i!=j:
                q[i, j] = mdl.addVar(vtype=GRB.BINARY, name="x%s,%s" % (i, j))
                mdl.update()

#for t in Cables_Index_Set:
#    print(t, type(t))


# Minimization Objective Function
if OWFCR_Type==0 or OWFCR_Type==1: # Equation 1 for OWFCR or SS
    mdl.setObjective(quicksum(Cable_Cost_Dict[t] * Distance_Dict[(i,j)] * x[i,j,t] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j for t in Cables_Index_Set),GRB.MINIMIZE)
elif OWFCR_Type==2: # Equation 16 for BP
    mdl.setObjective(quicksum(Cable_Cost_Dict[t] * Distance_Dict[(i,j)] * x[i,j,t] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j for t in Cables_Index_Set) + quicksum(pi_OWFCR_BR_Cost[d] * z[j,d] for d in D for j in Turbines_Index_Set),GRB.MINIMIZE)
elif OWFCR_Type==3: # Equation 24 for CL
    mdl.setObjective(quicksum(Cable_Cost_Dict[t] * Distance_Dict[(i,j)] * x[i,j,t] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j for t in Cables_Index_Set) + quicksum(min_cable_cost * Distance_Dict[(i,j)] * q[i,j] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j),GRB.MINIMIZE)
elif OWFCR_Type==4: # Equation 27 for OTM
    mdl.setObjective(quicksum(Cable_Cost_Dict[t] * Distance_Dict[(i,j)] * x[i,j,t] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j for t in Cables_Index_Set) + quicksum(c_otm * x[i,j,Tao] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j),GRB.MINIMIZE)
mdl.update()

# Equation 2
mdl.addConstrs(quicksum(x[i,j,t] for t in Cables_Index_Set) == y[i,j] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j)
mdl.update()

# Equation 3
mdl.addConstrs(quicksum(f[h,i] - f[i,h] for i in All_Nodes_Index_Set if i!=h) == All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[h][2] for h in (Turbines_Index_Set | Steiners_Index_Set))
mdl.update()

# Equation 4
mdl.addConstrs(quicksum(Cable_Capacity_Dict[t] * x[i,j,t] for t in Cables_Index_Set) >= f[i,j] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j)
mdl.update()

# Equation 5
mdl.addConstrs(quicksum(y[h,j] for j in All_Nodes_Index_Set if j!=h) == 1 for h in Turbines_Index_Set)
mdl.update()

# Equation 6
mdl.addConstrs(quicksum(y[h,j] for j in All_Nodes_Index_Set if j!=h) == 0 for h in Substations_Index_Set)
mdl.update()

# Equation 7
mdl.addConstrs(quicksum(y[h,j] for j in All_Nodes_Index_Set if j!=h) <= 1 for h in Steiners_Index_Set)
mdl.update()

# Equation 8
mdl.addConstrs(quicksum(y[i,h] for i in All_Nodes_Index_Set if i!=h) <= 1 for h in Steiners_Index_Set)
mdl.update()

# Equation 9
mdl.addConstrs(quicksum(y[i,h] for i in All_Nodes_Index_Set if i!=h) <= C for h in Substations_Index_Set)
mdl.update()


if No_Cross_Constraints == 0:
    # Equation 10
    for Each_Crossing in Crossings_Set:
        Line1 , Line2 = Each_Crossing
        h , k = Line1
        i , j = Line2
        mdl.addConstr(y[i,j] + y[j,i] + y[h,k] + y[k,h] <= 1)
        mdl.update()
elif No_Cross_Constraints == 1:
    # Equation 14
    mdl.addConstrs(quicksum(y[i,j] for i,j in Clique_Arc_Subset_Dict[(a,b,f)]) <= 1 for a in All_Nodes_Index_Set for b in All_Nodes_Index_Set if a!=b for f in All_Nodes_Index_Set if (f!=a and f!=b))
    mdl.update()

# Equation 15
if OWFCR_Type==1:
    mdl.addConstrs(quicksum(y[i,h] for i in All_Nodes_Index_Set if i!=h) <= 1 for h in Turbines_Index_Set)
    mdl.update()

# Equation 17 and 18
if OWFCR_Type==2:
    mdl.addConstrs(quicksum(y[i,j] for i in All_Nodes_Index_Set if i!=j) == quicksum(d * z[j,d] for d in D) for j in Turbines_Index_Set)
    mdl.update()
    
    mdl.addConstrs(quicksum(z[j,d] for d in D) <= 1 for j in Turbines_Index_Set)
    mdl.update()

#Equations 20, 21 and 23
if OWFCR_Type==3:
    mdl.addConstrs(quicksum(y[i,h]+y[h,i]+q[i,h]+q[h,i] for i in All_Nodes_Index_Set if i!=h) == 2 * quicksum(y[h,j] for j in All_Nodes_Index_Set if j!=h) for h in (Turbines_Index_Set | Steiners_Index_Set))
    mdl.update()
    
    mdl.addConstrs(q[i,j]==0 for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i>j)
    mdl.update()

    for Each_Crossing in Crossings_Set:
        Line1 , Line2 = Each_Crossing
        h , k = Line1
        i , j = Line2
        mdl.addConstr(y[i,j] + y[j,i] + y[h,k] + y[k,h] <= 1)
        mdl.addConstr(y[i,j] + y[j,i] + y[h,k] + y[k,h] + q[i,j] + q[j,i] + q[h,k] + q[k,h] <= 1)
        mdl.update()

# Equation 25 and 26
if OWFCR_Type==4:
    mdl.addConstrs(quicksum(x[i,h,t] for i in All_Nodes_Index_Set if i!=h for t in Cables_Index_Set if t!=Tao) <= mu_1 for h in Turbines_Index_Set)
    mdl.update()
    
    mdl.addConstrs(quicksum(x[i,h,Tao] for i in All_Nodes_Index_Set if i!=h) <= mu_2 for h in Turbines_Index_Set)
    mdl.update()


# Solve the Problem using default Gurobi settings
start_time=time()
mdl.optimize()
end_time=time()
winsound.Beep(500, 1000) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds

OptimalityGap=mdl.MIPGap
print("Final MIP Gap value: %f" % OptimalityGap)
best_bound=mdl.ObjBound

# run_time=mdl.Runtime
run_time=end_time-start_time
print("Best Bound found: ",best_bound," ; Found in ",run_time," seconds")

Solutions_Found=mdl.SolCount

if Solutions_Found:
    objec_val=mdl.getObjective().getValue() #mdl.objVal
    print("Objective Value: ",objec_val)


    if OWFCR_Type==3:
        rings_or_closed_loops=[]
        for t in Cables_Index_Set:
            if Cable_Cost_Dict[t]==min_cable_cost:
                for i in All_Nodes_Index_Set:
                    for j in All_Nodes_Index_Set:
                        if i!=j:
                            if (q[i,j].x<=1+tolerance) and (q[i,j].x>=1-tolerance):
                                rings_or_closed_loops.append((i,j))
                break


    All_Used_Steiner_Nodes=set()
    Cable_Routes_Dict={}

    # Plotting the Turbines, Substations, Steiner Nodes and Cable Routes in Separate Cable-Wise Map
    for t in Cables_Index_Set:
        routes=[]
        Used_Steiner_NodeSet=set()
        for i in All_Nodes_Index_Set:
            for j in All_Nodes_Index_Set:
                if i!=j:
                    if (x[i,j,t].x<=1+tolerance) and (x[i,j,t].x>=1-tolerance):
                        routes.append((i,j))
                        if i in Steiners_Index_Set:
                            Used_Steiner_NodeSet.add(i)
                        elif j in Steiners_Index_Set:
                            Used_Steiner_NodeSet.add(j)

        All_Used_Steiner_Nodes.union(Used_Steiner_NodeSet)
        Cable_Routes_Dict[t]=routes      

        """plt.figure(figsize=(11,11))
        
        for i in Used_Steiner_NodeSet:
            plt.scatter(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0], c='blue', marker="*")
            plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33, All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)

        for i in Turbines_Index_Set:
            plt.scatter(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1],All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0], c='r',marker='s')
            plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33, All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)

        for i in Substations_Index_Set:
            plt.scatter(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0], c='black')
            plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33, All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)


        for i,j in routes:
            #plt.annotate('', xy=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]], xytext=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]], arrowprops=dict(arrowstyle="-|>", connectionstyle='arc3', edgecolor=(counter/colour_intervals,1-(counter/colour_intervals),1)))
            plt.annotate('', xy=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]], xytext=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]], arrowprops=dict(arrowstyle="-|>", connectionstyle='arc3', edgecolor=(0.3,0.7,1)))

        plt.title("Cable Type "+str(t)+" having Capacity "+str(Cable_Capacity_Dict[t])+" and Costing "+str(Cable_Cost_Dict[t]))
        plt.ylabel("Latitude")
        plt.xlabel("Longitude")
        name="Off-Shore Wind Farm Cable Routes for "+str(t)+" type of Cable at Objective Value "+str(round(objec_val))+".jpg"
        plt.savefig(format(name),format='jpg',bbox_inches="tight")"""

    
    # Plotting the Turbines, Substations, Steiner Nodes and Cable Routes in Single Diagram
    plt.figure(figsize=(13,13))
    

    All_Used_Steiner_Nodes=Steiners_Index_Set
    Steiner_Handle=plt.scatter([All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] for i in All_Used_Steiner_Nodes], [All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] for i in All_Used_Steiner_Nodes], c='blue', marker="*", label="Steiner Nodes")
    handles.extend([Steiner_Handle])
    for i in All_Used_Steiner_Nodes:
        plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33, All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)


    # plot all turbines with the same marker and color, and assign the label outside the loop
    Turbine_Handle=plt.scatter([All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] for i in Turbines_Index_Set], [All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] for i in Turbines_Index_Set], c='r', marker='s', label="Turbines")
    handles.extend([Turbine_Handle])
    for i in Turbines_Index_Set:
        plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33, All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)

    
    Substation_Handle=plt.scatter([All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] for i in Substations_Index_Set], [All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] for i in Substations_Index_Set], c='black', label="Substations")
    handles.extend([Substation_Handle])
    for i in Substations_Index_Set:
        plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33, All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)


    Dynamic_Colour_Set=colourSet.copy()
    
    min_line_capacity=99999999999
    for t,capacity in Cable_Capacity_Dict.items():
        #print(t,capacity)
        if capacity<min_line_capacity:
            min_line_capacity=capacity

    for t, routes in Cable_Routes_Dict.items():        
        Colour_Me=random.choice(Dynamic_Colour_Set)
        Dynamic_Colour_Set.remove(Colour_Me)
        #print(Colour_Me,Dynamic_Colour_Set)
        dynamic_line_width = 1.1 * Cable_Capacity_Dict[t] / min_line_capacity

        if OWFCR_Type==3:
            if Cable_Cost_Dict[t]==min_cable_cost:
                for i,j in rings_or_closed_loops:
                    plt.annotate('', xy=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]], xytext=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]], arrowprops=dict(arrowstyle="<|-|>", connectionstyle='arc3', edgecolor=Colour_Me, linewidth=dynamic_line_width))
                    # https://matplotlib.org/stable/gallery/text_labels_and_annotations/fancyarrow_demo.html

        for i,j in routes:
            #plt.annotate('', xy=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]], xytext=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]], arrowprops=dict(arrowstyle="-|>", connectionstyle='arc3', edgecolor=(counter/colour_intervals,1-(counter/colour_intervals),1)))
            #plt.annotate('', xy=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]], xytext=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]], arrowprops=dict(arrowstyle="-|>", connectionstyle='arc3', edgecolor=(0.3,0.7,1)))
            #plt.annotate('', xy=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]], xytext=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]], arrowprops=dict(arrowstyle="-|>", connectionstyle='arc3', edgecolor=Colour_Me))
            plt.annotate('', xy=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]], xytext=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]], arrowprops=dict(arrowstyle="simple", connectionstyle='arc3', edgecolor=Colour_Me, linewidth=dynamic_line_width))
        
        line_handle = Line2D([0], [0], label=t, color=Colour_Me)
        #legend_item = Line2D([1000], [1000],marker='o',label="Point", color='r',markerfacecolor='k',markersize=10)
        #legend_i=plt.scatter([], [], c='red', marker="*", label='my marker')
        handles.extend([line_handle])

        if Dynamic_Colour_Set:
            continue
        else:
            Dynamic_Colour_Set=colourSet.copy()

    

    for individual_obstacle in Obstacles_Index_Set:
        obstacle_Lon, obstacle_Lat = [Obstacles_Line_Start_Lon_Dict[individual_obstacle],Obstacles_Line_End_Lon_Dict[individual_obstacle]], [Obstacles_Line_Start_Lat_Dict[individual_obstacle],Obstacles_Line_End_Lat_Dict[individual_obstacle]]
        #plt.plot(x1, y1, x2, y2, marker = 'o')
        plt.plot(obstacle_Lon, obstacle_Lat,linestyle="dashdot",color="black")
        mid_Lat = (Obstacles_Line_Start_Lat_Dict[individual_obstacle]+Obstacles_Line_End_Lat_Dict[individual_obstacle])/2
        mid_Lon = (Obstacles_Line_Start_Lon_Dict[individual_obstacle]+Obstacles_Line_End_Lon_Dict[individual_obstacle])/2
        plt.text(mid_Lon,mid_Lat,individual_obstacle)
        #plt.plot(x1, y1, x2, y2, marker = 'x')

    line_handle = Line2D([0], [0], linestyle="dashdot",label="Obstacles", color="black")
    handles.extend([line_handle])



    if No_Cross_Constraints == 0:
        no_cross=" & Eq 10"
    elif No_Cross_Constraints == 1:
        no_cross=" & Eq 14"
    else:
        no_cross=" without no-cross-constraints"
    solver_name="Gurobi_v10"+no_cross

    if OWFCR_Type==0:
        name="OWFCR Objective Value "+str(round(objec_val))+" kEuro found in "+str(round(run_time))+" seconds for "+str(len(Turbines_Index_Set))+" Turbines using "+solver_name
    elif OWFCR_Type==1:
        name="OWFCR_SS Objective Value "+str(round(objec_val))+" kEuro found in "+str(round(run_time))+" seconds for "+str(len(Turbines_Index_Set))+" Turbines using "+solver_name
    elif OWFCR_Type==2:
        name="OWFCR_BP Objective Value "+str(round(objec_val))+" kEuro found in "+str(round(run_time))+" seconds for "+str(len(Turbines_Index_Set))+" Turbines using "+solver_name
    elif OWFCR_Type==3:
        name="OWFCR_CL Objective Value "+str(round(objec_val))+" kEuro found in "+str(round(run_time))+" seconds for "+str(len(Turbines_Index_Set))+" Turbines using "+solver_name
    elif OWFCR_Type==4:
        name="OWFCR_OTM Objective Value "+str(round(objec_val))+" kEuro found in "+str(round(run_time))+" seconds for "+str(len(Turbines_Index_Set))+" Turbines using "+solver_name


    plt.title(name)
    plt.ylabel("Latitude")
    plt.xlabel("Longitude")
    plt.legend(handles=handles)
    #plt.legend()
    name=name+".png"
    plt.savefig(format(name),format='png',bbox_inches="tight")


    """
    # Call a Workbook() function of openpyxl to create a new blank Workbook object
    wb_individual = openpyxl.Workbook()
    # Get workbook active sheet from the active attribute
    sheet_individual = wb_individual.active
    row_number_on_Individual_Sheet=1
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
    cell.value = "From Node i"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
    cell.value = "To Node j"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
    cell.value = "y (indicating whether arc is built)"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
    cell.value = "f (representing the directed energy flow)"
    
    for col_number,t in enumerate(Cables_Index_Set):
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5 + col_number)
        cell.value = "Utilization of Cable Type "+str(t)

    if OWFCR_Type == 2:
        print("z[j,d] needs to be shown in the OutPut")
    elif OWFCR_Type == 3:
        print("q[i,j] needs to be shown in the OutPut")


    for i in All_Nodes_Index_Set:
        for j in All_Nodes_Index_Set:
            if i!=j:
                row_number_on_Individual_Sheet+=1

                cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
                cell.value = i
                cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
                cell.value = j
                cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
                cell.value = y[i,j].x
                cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
                cell.value = f[i,j].x
                
                for col_number,t in enumerate(Cables_Index_Set):
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5 + col_number)
                    cell.value = x[i,j,t].x

    #wb_individual.save(str(directory_to_save_Gurobi_solution)+"Solution Details.xlsx")
    wb_individual.save("Solution Details.xlsx")
    """

else:
    print("NO SOLUTION FOUND")
