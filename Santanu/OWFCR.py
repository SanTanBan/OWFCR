import winsound
import openpyxl
from math import sqrt
import matplotlib.pylab as plt
import pandas as pd
from gurobipy import *
tolerance=1e-5 # This is the default value for Gurobi

# An additional technical requirement is that a given maximum number of cables, say C, can be connected to each substation.
C = 3



def Euclidean_Distance_Matrix_Generator(locations_dict, output_filename):
    Distance_Dict = {}
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Euclidean Distances")

    # Write headers
    ws.cell(row=1, column=1, value="Location")
    for i, loc in enumerate(locations_dict):
        #print(i,loc)
        ws.cell(row=1, column=i+2, value=loc)
    
    # Write distance matrix
    for i, loc1 in enumerate(locations_dict):
        ws.cell(row=i+2, column=1, value=loc1)
        for j, loc2 in enumerate(locations_dict):
            if i == j:
                ws.cell(row=i+2, column=j+2, value=0)
                Distance_Dict[(loc1,loc2)] = 0
            elif j < i:
                ws.cell(row=i+2, column=j+2, value=ws.cell(row=j+2, column=i+2).value)
                Distance_Dict[(loc1,loc2)] = Distance_Dict[(loc2,loc1)]
            else:
                lat1, lon1, pow1 = locations_dict[loc1]
                lat2, lon2, pow2 = locations_dict[loc2]
                #distance = round(sqrt((lat1-lat2)**2 + (lon1-lon2)**2),2)
                distance = sqrt((lat1-lat2)**2 + (lon1-lon2)**2)
                ws.cell(row=i+2, column=j+2, value=distance)
                Distance_Dict[(loc1,loc2)] = distance
    
    del wb["Sheet"]
    # Save workbook to file
    wb.save(output_filename)
    return Distance_Dict



def do_lines_intersect(line1_start, line1_end, line2_start, line2_end):
    """
    Check if two line segments intersect.

    Args:
        line1_start (tuple): The starting point of the first line segment, represented as a tuple of (x, y) coordinates.
        line1_end (tuple): The ending point of the first line segment, represented as a tuple of (x, y) coordinates.
        line2_start (tuple): The starting point of the second line segment, represented as a tuple of (x, y) coordinates.
        line2_end (tuple): The ending point of the second line segment, represented as a tuple of (x, y) coordinates.

    Returns:
        bool: True if the two line segments intersect, False otherwise.
    """
    x1, y1 , p1 = line1_start
    x2, y2 , p2 = line1_end
    x3, y3 , p3 = line2_start
    x4, y4 , p4 = line2_end

    # Calculate the slopes of the two lines
    slope1 = (y2 - y1) / (x2 - x1) if x2 - x1 != 0 else float('inf')
    slope2 = (y4 - y3) / (x4 - x3) if x4 - x3 != 0 else float('inf')

    # Check if the lines are parallel
    if slope1 == slope2:
        return False

    # Calculate the y-intercepts of the two lines
    intercept1 = y1 - slope1 * x1
    intercept2 = y3 - slope2 * x3

    # Calculate the x-coordinate of the point of intersection
    x_intersect = (intercept2 - intercept1) / (slope1 - slope2)

    # Check if the x-coordinate of the point of intersection is within the bounds of both line segments
    if x_intersect < min(x1, x2) or x_intersect > max(x1, x2) or x_intersect < min(x3, x4) or x_intersect > max(x3, x4):
        return False

    return True



def Intersecting_Arcs_Combo(locations_dict, output_filename):

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Crossing Segments")
    row_count=1
    ws.cell(row=row_count, column=1, value="Line 1 Start")
    ws.cell(row=row_count, column=2, value="Line 1 End")
    ws.cell(row=row_count, column=3, value="Line 2 Start")
    ws.cell(row=row_count, column=4, value="Line 2 End")

    Crossings_Set=set()

    for i, start_line1 in enumerate(locations_dict):
        for j, start_line2 in enumerate(locations_dict):
            #if start_line1!=start_line2 and i<j:
            if i<j:
                for k, end_line1 in enumerate(locations_dict):
                    #if end_line1!=start_line2 and start_line1!=end_line1 and i<k:
                    if end_line1!=start_line2 and i<k:
                        for l, end_line2 in enumerate(locations_dict):
                            #if end_line2!=end_line1 and end_line2!=start_line2 and end_line2!=start_line1 and j<l:
                            if end_line2!=end_line1 and j<l:
                                if do_lines_intersect(locations_dict[start_line1], locations_dict[end_line1], locations_dict[start_line2], locations_dict[end_line2]):
                                    #Crossings_Set.add(frozenset([frozenset([start_line1, end_line1]),frozenset([start_line2, end_line2])]))
                                    Crossings_Set.add(((start_line1, end_line1),(start_line2, end_line2)))
                                    row_count+=1
                                    ws.cell(row=row_count, column=1, value=start_line1)
                                    ws.cell(row=row_count, column=2, value=end_line1)
                                    ws.cell(row=row_count, column=3, value=start_line2)
                                    ws.cell(row=row_count, column=4, value=end_line2)
                                else:
                                    continue

    del wb["Sheet"]
    # Save workbook to file
    wb.save(output_filename)
    return Crossings_Set



Turbines=pd.read_excel("Input.xlsx","Turbine Locations",index_col=0)
Substations=pd.read_excel("Input.xlsx","Substation Locations",index_col=0)
Steiners=pd.read_excel("Input.xlsx","Steiner Locations",index_col=0)
Cables=pd.read_excel("Input.xlsx","Cable Specifications",index_col=0)

# print(Turbines)
# print(Substations)
# print(Steiners)
# print(Cables)

Cables_Index_Set = set(Cables.index)
Cable_Cost_Dict = Cables["Cost"].to_dict()
Cable_Capacity_Dict = Cables["Capacity"].to_dict()

Turbines_Index_Set = set(Turbines.index)
Turbines_Latitude_Dict = Turbines["Latitude"].to_dict()
Turbines_Longitude_Dict = Turbines["Longitude"].to_dict()
Turbines_Power_Dict = Turbines["Power"].to_dict()

Substations_Index_Set = set(Substations.index)
Substations_Latitude_Dict = Substations["Latitude"].to_dict()
Substations_Longitude_Dict = Substations["Longitude"].to_dict()
Substations_Power_Dict = Substations["Power"].to_dict()

Steiners_Index_Set = set(Steiners.index)
Steiners_Latitude_Dict = Steiners["Latitude"].to_dict()
Steiners_Longitude_Dict = Steiners["Longitude"].to_dict()
Steiners_Power_Dict = Steiners["Power"].to_dict()

All_Nodes_Index_Set = Turbines_Index_Set | Substations_Index_Set | Steiners_Index_Set
All_Nodes_DictKey_Lat_Lon_Power_ValueTuple={}
for Single_Node in All_Nodes_Index_Set:
    if Single_Node in Turbines_Index_Set:
        ValueTuple=(Turbines_Latitude_Dict[Single_Node],Turbines_Longitude_Dict[Single_Node],Turbines_Power_Dict[Single_Node])
    elif Single_Node in Substations_Index_Set:
        ValueTuple=(Substations_Latitude_Dict[Single_Node],Substations_Longitude_Dict[Single_Node],Substations_Power_Dict[Single_Node])
    elif Single_Node in Steiners_Index_Set:
        ValueTuple=(Steiners_Latitude_Dict[Single_Node],Steiners_Latitude_Dict[Single_Node],Steiners_Power_Dict[Single_Node])
    else:
        print("Fatal Error! ",Single_Node," not found anywhere")
    All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[Single_Node]=ValueTuple

# print(All_Nodes_Index_Set)
# print(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple)

Distance_Dict = Euclidean_Distance_Matrix_Generator(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple, "Distance Matrix.xlsx")
# print(Distance_Dict)

Crossings_Set=Intersecting_Arcs_Combo(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple, "Crossing Segments.xlsx")
print("Number of Crossings found: ",len(Crossings_Set))

# for Each_Crossing in Crossings_Set:
#     Line1 , Line2 = Each_Crossing
#     h , k = Line1
#     i , j = Line2
#     print(h,k,i,j)


# Set the problem
mdl=Model("VRP")
mdl.setParam('TimeLimit', 9999)

f, x, y = {}, {}, {}
# Decision Variables

for i in All_Nodes_Index_Set:
    for j in All_Nodes_Index_Set:
        if i!=j:
            f[i, j] = mdl.addVar(vtype=GRB.CONTINUOUS, name="x%s,%s" % (i, j))
            mdl.update()

for i in All_Nodes_Index_Set:
    for j in All_Nodes_Index_Set:
        if i!=j:
            for t in Cables_Index_Set:
                x[i, j, t] = mdl.addVar(vtype=GRB.BINARY, name="x%s,%s,%s" % (i, j, t))
                mdl.update()

for i in All_Nodes_Index_Set:
    for j in All_Nodes_Index_Set:
        if i!=j:
            y[i, j] = mdl.addVar(vtype=GRB.BINARY, name="x%s,%s" % (i, j))
            mdl.update()


# Equation 1
mdl.setObjective(quicksum(Cable_Cost_Dict[t] * Distance_Dict[(i,j)] * x[i,j,t] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j for t in Cables_Index_Set),GRB.MINIMIZE)
mdl.update()
# Equation 2
mdl.addConstrs(quicksum(x[i,j,t] for t in Cables_Index_Set) == y[i,j] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j)
mdl.update()
# Equation 3
mdl.addConstrs(quicksum(f[h,i] - f[i,h] for i in All_Nodes_Index_Set if i!=h) == All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[h][2] for h in (Turbines_Index_Set | Steiners_Index_Set))
mdl.update()
# Equation 4
mdl.addConstrs(quicksum(Cable_Capacity_Dict[t] * x[i,j,t] for t in Cables_Index_Set) >= f[i,j] for i in All_Nodes_Index_Set for j in All_Nodes_Index_Set if i!=j)
mdl.update()
# Equation 5
mdl.addConstrs(quicksum(y[h,j] for j in All_Nodes_Index_Set if j!=h) == 1 for h in Turbines_Index_Set)
mdl.update()
# Equation 6
mdl.addConstrs(quicksum(y[h,j] for j in All_Nodes_Index_Set if j!=h) == 0 for h in Substations_Index_Set)
mdl.update()
# Equation 7
mdl.addConstrs(quicksum(y[h,j] for j in All_Nodes_Index_Set if j!=h) <= 1 for h in Steiners_Index_Set)
mdl.update()
# Equation 8
mdl.addConstrs(quicksum(y[i,h] for i in All_Nodes_Index_Set if i!=h) <= 1 for h in Steiners_Index_Set)
mdl.update()
# Equation 9
mdl.addConstrs(quicksum(y[i,h] for i in All_Nodes_Index_Set if i!=h) <= C for h in Substations_Index_Set)
mdl.update()
# Equation 10
for Each_Crossing in Crossings_Set:
    Line1 , Line2 = Each_Crossing
    h , k = Line1
    i , j = Line2
    mdl.addConstr(y[i,j] + y[j,i] + y[h,k] + y[k,h] <= 1)
    mdl.update()


# Solve the Problem using default Gurobi settings
mdl.optimize()
winsound.Beep(500, 1000) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds

OptimalityGap=mdl.MIPGap
print("Final MIP Gap value: %f" % OptimalityGap)
best_bound=mdl.ObjBound
print("Best Bound found: ",best_bound)

Solutions_Found=mdl.SolCount

if Solutions_Found:
    objec_val=mdl.getObjective().getValue() #mdl.objVal
    print(objec_val)




    # Plotting the Turbines, Substations, Steiner Nodes and Cable Routes
    for t in Cables_Index_Set:
        plt.figure(figsize=(11,11))
        

        routes=[]
        Used_Steiner_NodeSet=set()
        for i in All_Nodes_Index_Set:
            for j in All_Nodes_Index_Set:
                if i!=j:
                    if (x[i,j,t].x<=1+tolerance) and (x[i,j,t].x>=1-tolerance):
                        routes.append((i,j))
                        if i in Steiners_Index_Set:
                            Used_Steiner_NodeSet.add(i)
                        elif j in Steiners_Index_Set:
                            Used_Steiner_NodeSet.add(j)


        for i in Used_Steiner_NodeSet:
            plt.scatter(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0], c='blue', marker="*")
            plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33, All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)

        for i in Turbines_Index_Set:
            plt.scatter(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1],All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0], c='r',marker='s')
            plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33, All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)

        for i in Substations_Index_Set:
            plt.scatter(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0], c='black')
            plt.text(All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1] + 0.33, All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0] + 0.33, i)


        for i,j in routes:
            #plt.annotate('', xy=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]], xytext=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]], arrowprops=dict(arrowstyle="-|>", connectionstyle='arc3', edgecolor=(counter/colour_intervals,1-(counter/colour_intervals),1)))
            plt.annotate('', xy=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[j][0]], xytext=[All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][1], All_Nodes_DictKey_Lat_Lon_Power_ValueTuple[i][0]], arrowprops=dict(arrowstyle="-|>", connectionstyle='arc3', edgecolor=(0.3,0.7,1)))

        plt.title("Cable Type "+str(t)+" having Capacity "+str(Cable_Capacity_Dict[t])+" and Costing "+str(Cable_Cost_Dict[t]))
        plt.ylabel("Latitude")
        plt.xlabel("Longitude")

        
        
        name="Off-Shore Wind Farm Cable Routes for "+str(t)+" type of Cable at Objective Value "+str(round(objec_val))+".jpg"
        plt.savefig(format(name),format='jpg',bbox_inches="tight")


    

    # Call a Workbook() function of openpyxl to create a new blank Workbook object
    wb_individual = openpyxl.Workbook()
    # Get workbook active sheet from the active attribute
    sheet_individual = wb_individual.active
    row_number_on_Individual_Sheet=1
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
    cell.value = "From Node i"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
    cell.value = "To Node j"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
    cell.value = "y (indicating whether arc is built)"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
    cell.value = "f (representing the directed energy flow)"
    
    for col_number,t in enumerate(Cables_Index_Set):
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5 + col_number)
        cell.value = "Utilization of Cable Type "+str(t)


    for i in All_Nodes_Index_Set:
        for j in All_Nodes_Index_Set:
            if i!=j:
                row_number_on_Individual_Sheet+=1

                cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
                cell.value = i
                cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
                cell.value = j
                cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
                cell.value = y[i,j].x
                cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
                cell.value = f[i,j].x
                
                for col_number,t in enumerate(Cables_Index_Set):
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5 + col_number)
                    cell.value = x[i,j,t].x

    #wb_individual.save(str(directory_to_save_Gurobi_solution)+"Solution Details.xlsx")
    wb_individual.save("Solution Details.xlsx")
